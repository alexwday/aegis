"""
WM Readthrough ETL Script - Template-Driven Section Processing.

This refactored script uses a template-driven approach where sections are defined
in an external CSV template with configurable prompts and bank filters.

Template Structure:
- section_id: Unique identifier for the section
- section_name: Human-readable section name
- section_description: What to extract
- section_instructions: Specific extraction instructions
- section_notes: Additional guidance
- section_examples: Example outputs
- transcript_parts: MD, QA, or BOTH
- institution_types: Comma-separated bank types
- prompt_name: Which postgres prompt template to use

Usage:
    python -m aegis.etls.wm_readthrough.main_refactored --year 2025 --quarter Q1
    python -m aegis.etls.wm_readthrough.main_refactored --year 2025 --quarter Q1 \
        --template custom_sections.csv
"""

import argparse
import asyncio
import csv
import hashlib
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

from sqlalchemy import text

# Import direct transcript functions
from aegis.model.subagents.transcripts.retrieval import retrieve_full_section
from aegis.model.subagents.transcripts.formatting import format_full_section_chunks
from aegis.utils.ssl import setup_ssl
from aegis.connections.oauth_connector import setup_authentication
from aegis.connections.llm_connector import complete_with_tools
from aegis.connections.postgres_connector import get_connection
from aegis.utils.logging import setup_logging, get_logger
from aegis.utils.prompt_loader import load_prompt_from_db
from aegis.etls.wm_readthrough.config.config import (
    MODELS,
    TEMPERATURE,
    MAX_TOKENS,
    MAX_CONCURRENT_BANKS,
    SECTION_TEMPLATE_PATH,
    get_monitored_institutions,
)

# Initialize logging
setup_logging()
logger = get_logger()

# =============================================================================
# TEMPLATE LOADING
# =============================================================================


def load_section_template(template_path: str) -> List[Dict[str, Any]]:
    """
    Load section definitions from CSV or XLSX template.

    Args:
        template_path: Path to CSV or XLSX template file

    Returns:
        List of section definition dictionaries

    Expected columns:
        - section_id
        - section_name
        - section_description
        - section_instructions
        - section_notes
        - section_examples
        - transcript_parts (MD, QA, or ALL)
        - institution_types (comma-separated)
        - prompt_name
    """
    template_path_obj = Path(template_path)
    if not template_path_obj.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    sections = []
    file_extension = template_path_obj.suffix.lower()

    if file_extension == ".xlsx":
        # Load XLSX file
        from openpyxl import load_workbook

        wb = load_workbook(template_path, read_only=True, data_only=True)
        ws = wb.active

        # Get header row
        headers = [cell.value for cell in ws[1]]

        # Process data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            row_dict = dict(zip(headers, row))

            # Parse institution types
            institution_types = [
                t.strip() for t in str(row_dict["institution_types"]).split(",")
            ]

            sections.append(
                {
                    "section_id": str(row_dict["section_id"]).strip(),
                    "section_name": str(row_dict["section_name"]).strip(),
                    "section_description": str(row_dict["section_description"]).strip(),
                    "section_instructions": str(row_dict["section_instructions"]).strip(),
                    "section_notes": str(row_dict["section_notes"]).strip(),
                    "section_examples": str(row_dict["section_examples"]).strip(),
                    "transcript_parts": str(row_dict["transcript_parts"]).strip().upper(),
                    "institution_types": institution_types,
                    "prompt_name": str(row_dict["prompt_name"]).strip(),
                }
            )

        wb.close()

    elif file_extension == ".csv":
        # Load CSV file
        with open(template_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Parse institution types
                institution_types = [t.strip() for t in row["institution_types"].split(",")]

                sections.append(
                    {
                        "section_id": row["section_id"].strip(),
                        "section_name": row["section_name"].strip(),
                        "section_description": row["section_description"].strip(),
                        "section_instructions": row["section_instructions"].strip(),
                        "section_notes": row["section_notes"].strip(),
                        "section_examples": row["section_examples"].strip(),
                        "transcript_parts": row["transcript_parts"].strip().upper(),
                        "institution_types": institution_types,
                        "prompt_name": row["prompt_name"].strip(),
                    }
                )
    else:
        raise ValueError(f"Unsupported template format: {file_extension}. Use .csv or .xlsx")

    logger.info(f"[TEMPLATE] Loaded {len(sections)} section definitions from {template_path}")
    return sections


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


async def get_bank_info(bank_identifier: Any) -> Dict[str, Any]:
    """
    Resolve bank identifier (ID, symbol, or name) to full bank information.

    Args:
        bank_identifier: Bank ID (int), symbol (str), or name (str)

    Returns:
        Dict with bank_id, bank_name, and bank_symbol
    """
    async with get_connection() as conn:
        # Try different identifier types
        if isinstance(bank_identifier, int) or (
            isinstance(bank_identifier, str) and bank_identifier.isdigit()
        ):
            # Bank ID provided
            query = text(
                """
                SELECT DISTINCT bank_id, bank_name, bank_symbol
                FROM aegis_data_availability
                WHERE bank_id = :bank_id
                LIMIT 1
            """
            )
            result = await conn.execute(query, {"bank_id": int(bank_identifier)})
        else:
            # Try symbol or name
            query = text(
                """
                SELECT DISTINCT bank_id, bank_name, bank_symbol
                FROM aegis_data_availability
                WHERE UPPER(bank_symbol) = UPPER(:identifier)
                   OR UPPER(bank_name) = UPPER(:identifier)
                LIMIT 1
            """
            )
            result = await conn.execute(query, {"identifier": str(bank_identifier)})

        row = result.first()
        if not row:
            raise ValueError(f"Bank not found: {bank_identifier}")

        return {"bank_id": row.bank_id, "bank_name": row.bank_name, "bank_symbol": row.bank_symbol}


async def find_latest_available_quarter(
    bank_id: int, min_fiscal_year: int, min_quarter: str, bank_name: str = ""
) -> Optional[Tuple[int, str]]:
    """
    Find the latest available quarter for a bank, at or after the minimum specified.

    Args:
        bank_id: Bank ID
        min_fiscal_year: Minimum fiscal year
        min_quarter: Minimum quarter
        bank_name: Bank name for logging

    Returns:
        Tuple of (fiscal_year, quarter) if found, None otherwise
    """
    async with get_connection() as conn:
        # Convert quarter to sortable format
        quarter_map = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
        min_quarter_num = quarter_map.get(min_quarter, 1)

        query = text(
            """
            SELECT fiscal_year, quarter
            FROM aegis_data_availability
            WHERE bank_id = :bank_id
              AND 'transcripts' = ANY(database_names)
              AND (fiscal_year > :min_year
                   OR (fiscal_year = :min_year
                       AND CASE quarter
                           WHEN 'Q1' THEN 1
                           WHEN 'Q2' THEN 2
                           WHEN 'Q3' THEN 3
                           WHEN 'Q4' THEN 4
                       END >= :min_quarter))
            ORDER BY fiscal_year DESC,
                     CASE quarter
                         WHEN 'Q4' THEN 4
                         WHEN 'Q3' THEN 3
                         WHEN 'Q2' THEN 2
                         WHEN 'Q1' THEN 1
                     END DESC
            LIMIT 1
        """
        )

        result = await conn.execute(
            query, {"bank_id": bank_id, "min_year": min_fiscal_year, "min_quarter": min_quarter_num}
        )

        row = result.first()
        if row:
            latest_year = row.fiscal_year
            latest_quarter = row.quarter

            # Log if we're using a more recent quarter
            if latest_year > min_fiscal_year or (
                latest_year == min_fiscal_year
                and quarter_map.get(latest_quarter, 0) > min_quarter_num
            ):
                logger.info(
                    f"[LATEST QUARTER] {bank_name or f'Bank {bank_id}'}: "
                    f"Using {latest_year} {latest_quarter} "
                    f"(requested minimum: {min_fiscal_year} {min_quarter})"
                )

            return (latest_year, latest_quarter)

        return None


async def retrieve_transcript_parts(
    bank_info: Dict[str, Any],
    fiscal_year: int,
    quarter: str,
    context: Dict[str, Any],
    parts: str,
    use_latest: bool,
) -> Optional[str]:
    """
    Retrieve specified transcript parts (MD, QA, or ALL).

    Args:
        bank_info: Bank information dictionary
        fiscal_year: Fiscal year
        quarter: Quarter
        context: Execution context
        parts: "MD", "QA", or "ALL"
        use_latest: Whether to use latest available quarter

    Returns:
        Combined transcript string or None if not available
    """
    # Determine which quarter to use
    actual_year, actual_quarter = fiscal_year, quarter

    if use_latest:
        latest = await find_latest_available_quarter(
            bank_id=bank_info["bank_id"],
            min_fiscal_year=fiscal_year,
            min_quarter=quarter,
            bank_name=bank_info["bank_name"],
        )
        if latest:
            actual_year, actual_quarter = latest
        else:
            logger.warning(
                f"[NO DATA] {bank_info['bank_name']}: "
                f"No transcript data for {fiscal_year} {quarter} or later"
            )
            return None

    # Build combo dict for transcript retrieval
    combo = {
        "bank_id": bank_info["bank_id"],
        "bank_name": bank_info["bank_name"],
        "bank_symbol": bank_info["bank_symbol"],
        "fiscal_year": actual_year,
        "quarter": actual_quarter,
    }

    try:
        content_parts = []

        # Get MD section if needed
        if parts in ["MD", "ALL"]:
            md_chunks = await retrieve_full_section(combo=combo, sections="MD", context=context)
            md_content = await format_full_section_chunks(
                chunks=md_chunks, combo=combo, context=context
            )
            content_parts.append(md_content)

        # Get Q&A section if needed
        if parts in ["QA", "ALL"]:
            qa_chunks = await retrieve_full_section(combo=combo, sections="QA", context=context)
            qa_content = await format_full_section_chunks(
                chunks=qa_chunks, combo=combo, context=context
            )
            content_parts.append(qa_content)

        combined = "\n\n".join(content_parts)

        logger.info(
            f"[TRANSCRIPT] {bank_info['bank_name']} {actual_year} {actual_quarter}: "
            f"Retrieved {len(combined)} chars ({parts} sections)"
        )

        return combined

    except Exception as e:
        logger.error(f"Error retrieving transcript for {bank_info['bank_name']}: {e}")
        return None


# =============================================================================
# GENERIC SECTION EXTRACTION
# =============================================================================


async def extract_section(
    section_def: Dict[str, Any],
    bank_info: Dict[str, Any],
    transcript_content: str,
    fiscal_year: int,
    quarter: str,
    context: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Generic section extraction using template-driven prompts.

    Args:
        section_def: Section definition from template
        bank_info: Bank information dictionary
        transcript_content: Transcript text
        fiscal_year: Year
        quarter: Quarter
        context: Execution context

    Returns:
        {
            "section_id": str,
            "section_name": str,
            "bank_name": str,
            "bank_symbol": str,
            "has_content": bool,
            "content": str,
            "metadata": dict (optional)
        }
    """
    execution_id = context.get("execution_id")

    # Load prompt template from database
    prompt_template = load_prompt_from_db(
        layer="wm_readthrough_etl",
        name=section_def["prompt_name"],
        compose_with_globals=False,
        execution_id=execution_id,
    )

    # Format system prompt with section variables
    system_prompt = prompt_template["system_prompt"].format(
        section_description=section_def["section_description"],
        section_instructions=section_def["section_instructions"],
        section_notes=section_def["section_notes"],
        section_examples=section_def["section_examples"],
    )

    # Format user prompt with all variables
    user_prompt = prompt_template["user_prompt"].format(
        bank_name=bank_info["bank_name"],
        fiscal_year=fiscal_year,
        quarter=quarter,
        section_name=section_def["section_name"],
        transcript_content=transcript_content,
    )

    # Build messages
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    # Create tool definition
    tool_def = prompt_template.get("tool_definition")
    if tool_def:
        tools = [tool_def]
    else:
        logger.warning(f"No tool definition found for {section_def['prompt_name']}")
        return {
            "section_id": section_def["section_id"],
            "section_name": section_def["section_name"],
            "bank_name": bank_info["bank_name"],
            "bank_symbol": bank_info["bank_symbol"],
            "has_content": False,
            "content": "",
            "error": "No tool definition in prompt",
        }

    # Call LLM
    llm_params = {
        "model": MODELS.get("page1_wm_narrative", "gpt-4-turbo"),  # Use default model
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
    }

    try:
        response = await complete_with_tools(
            messages=messages, tools=tools, context=context, llm_params=llm_params
        )

        # Extract tool call results
        tool_calls = response.get("choices", [{}])[0].get("message", {}).get("tool_calls")
        if tool_calls:
            tool_call = tool_calls[0]
            result = json.loads(tool_call["function"]["arguments"])

            if not result.get("has_content", False):
                logger.info(
                    f"[NO CONTENT] {section_def['section_id']} - {bank_info['bank_name']}: "
                    f"No relevant content found"
                )
                return {
                    "section_id": section_def["section_id"],
                    "section_name": section_def["section_name"],
                    "bank_name": bank_info["bank_name"],
                    "bank_symbol": bank_info["bank_symbol"],
                    "has_content": False,
                    "content": "",
                }

            logger.info(
                f"[EXTRACTED] {section_def['section_id']} - {bank_info['bank_name']}: "
                f"{len(result.get('content', ''))} chars"
            )

            return {
                "section_id": section_def["section_id"],
                "section_name": section_def["section_name"],
                "bank_name": bank_info["bank_name"],
                "bank_symbol": bank_info["bank_symbol"],
                "has_content": True,
                "content": result.get("content", ""),
                "metadata": result.get("metadata", {}),
            }
        else:
            logger.warning(
                f"No tool call in response for "
                f"{section_def['section_id']} - {bank_info['bank_name']}"
            )
            return {
                "section_id": section_def["section_id"],
                "section_name": section_def["section_name"],
                "bank_name": bank_info["bank_name"],
                "bank_symbol": bank_info["bank_symbol"],
                "has_content": False,
                "content": "",
            }

    except Exception as e:
        logger.error(
            f"Error extracting {section_def['section_id']} for {bank_info['bank_name']}: {e}"
        )
        return {
            "section_id": section_def["section_id"],
            "section_name": section_def["section_name"],
            "bank_name": bank_info["bank_name"],
            "bank_symbol": bank_info["bank_symbol"],
            "has_content": False,
            "content": "",
            "error": str(e),
        }


# =============================================================================
# MAIN ORCHESTRATION
# =============================================================================


async def process_all_sections(
    section_defs: List[Dict[str, Any]],
    fiscal_year: int,
    quarter: str,
    context: Dict[str, Any],
    use_latest: bool = False,
) -> List[Dict[str, Any]]:
    """
    Process all sections for all applicable banks using template definitions.

    Args:
        section_defs: List of section definitions from template
        fiscal_year: Year
        quarter: Quarter
        context: Execution context
        use_latest: If True, use latest available quarter >= specified quarter

    Returns:
        List of extraction results
    """
    # Load all institutions
    all_institutions = get_monitored_institutions()

    logger.info(
        f"[START] Processing {fiscal_year} {quarter} | "
        f"{len(section_defs)} sections | "
        f"{len(all_institutions)} total institutions | "
        f"Mode: {'latest available' if use_latest else 'exact quarter'}"
    )

    # Concurrency control
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_BANKS)

    # Process each section
    all_results = []

    for section_def in section_defs:
        logger.info(f"\n[SECTION START] {section_def['section_id']}: {section_def['section_name']}")

        # Filter banks for this section
        applicable_banks = [
            bank
            for bank in all_institutions
            if bank.get("type") in section_def["institution_types"]
        ]

        logger.info(
            f"[SECTION BANKS] {section_def['section_id']}: "
            f"{len(applicable_banks)} banks matched (types: {section_def['institution_types']})"
        )

        # Process banks for this section
        async def process_bank_section(bank_data, current_section=section_def):
            async with semaphore:
                try:
                    bank_info = await get_bank_info(bank_data["bank_id"])

                    # Retrieve transcript parts
                    transcript = await retrieve_transcript_parts(
                        bank_info,
                        fiscal_year,
                        quarter,
                        context,
                        current_section["transcript_parts"],
                        use_latest,
                    )

                    if not transcript:
                        return {
                            "section_id": current_section["section_id"],
                            "section_name": current_section["section_name"],
                            "bank_name": bank_info["bank_name"],
                            "bank_symbol": bank_info["bank_symbol"],
                            "has_content": False,
                            "content": "",
                        }

                    # Extract section content
                    result = await extract_section(
                        current_section, bank_info, transcript, fiscal_year, quarter, context
                    )

                    return result

                except Exception as e:
                    logger.error(
                        f"Error processing {current_section['section_id']} for {bank_data}: {e}"
                    )
                    return {
                        "section_id": current_section["section_id"],
                        "section_name": current_section["section_name"],
                        "bank_name": bank_data.get("bank_name", "Unknown"),
                        "bank_symbol": bank_data.get("bank_symbol", ""),
                        "has_content": False,
                        "content": "",
                        "error": str(e),
                    }

        # Execute all banks for this section concurrently
        tasks = [process_bank_section(bank, section_def) for bank in applicable_banks]
        section_results = await asyncio.gather(*tasks, return_exceptions=True)

        # Filter out exceptions and add to results
        for result in section_results:
            if not isinstance(result, Exception):
                all_results.append(result)

        # Log section completion
        content_count = sum(
            1 for r in section_results if not isinstance(r, Exception) and r.get("has_content")
        )
        logger.info(
            f"[SECTION COMPLETE] {section_def['section_id']}: "
            f"{content_count}/{len(applicable_banks)} banks with content"
        )

    logger.info(
        f"\n[PIPELINE COMPLETE] All sections processed | "
        f"Total results: {len(all_results)} | "
        f"With content: {sum(1 for r in all_results if r.get('has_content'))}"
    )

    return all_results


# =============================================================================
# DATABASE STORAGE
# =============================================================================


async def save_to_database(
    results: List[Dict[str, Any]],
    fiscal_year: int,
    quarter: str,
    execution_id: str,
    output_json_path: str = None,
) -> None:
    """
    Save extraction results to database and optionally to JSON file.

    Args:
        results: List of extraction results
        fiscal_year: Year
        quarter: Quarter
        execution_id: Execution UUID
        output_json_path: Optional path to save JSON output
    """
    # Save to JSON file if requested
    if output_json_path:
        output_data = {
            "metadata": {
                "fiscal_year": fiscal_year,
                "quarter": quarter,
                "total_extractions": len(results),
                "extractions_with_content": sum(1 for r in results if r.get("has_content")),
                "generation_date": datetime.now().isoformat(),
                "execution_id": str(execution_id),
            },
            "results": results,
        }

        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)

        logger.info(f"[OUTPUT] Results saved to {output_json_path}")

    # Save to database
    async with get_connection() as conn:
        query = text(
            """
            INSERT INTO aegis_reports (
                report_name,
                report_description,
                report_type,
                bank_id,
                bank_name,
                bank_symbol,
                fiscal_year,
                quarter,
                markdown_content,
                generation_date,
                generated_by,
                execution_id,
                metadata
            ) VALUES (
                :report_name,
                :report_description,
                :report_type,
                :bank_id,
                :bank_name,
                :bank_symbol,
                :fiscal_year,
                :quarter,
                :markdown_content,
                NOW(),
                :generated_by,
                :execution_id,
                :metadata
            )
        """
        )

        # Create markdown summary
        markdown_lines = [
            f"# WM Readthrough Report - {fiscal_year} {quarter}",
            "",
            "## Summary",
            f"- Total Extractions: {len(results)}",
            f"- With Content: {sum(1 for r in results if r.get('has_content'))}",
            "",
        ]

        await conn.execute(
            query,
            {
                "report_name": "WM Readthrough (Template-Driven)",
                "report_description": (
                    "AI-generated WM analysis using template-driven section processing"
                ),
                "report_type": "wm_readthrough_v2",
                "bank_id": None,
                "bank_name": None,
                "bank_symbol": None,
                "fiscal_year": fiscal_year,
                "quarter": quarter,
                "markdown_content": "\n".join(markdown_lines),
                "generated_by": "wm_readthrough_etl_v2",
                "execution_id": str(execution_id),
                "metadata": json.dumps({"results": results}),
            },
        )

        await conn.commit()

    logger.info(f"[DATABASE] Report saved with execution_id: {execution_id}")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================


async def main():
    """Main entry point for the template-driven WM Readthrough ETL."""
    parser = argparse.ArgumentParser(
        description="Generate WM Readthrough report using template-driven section processing"
    )
    parser.add_argument("--year", type=int, required=True, help="Fiscal year (e.g., 2025)")
    parser.add_argument("--quarter", type=str, required=True, help="Quarter (e.g., Q1)")
    parser.add_argument(
        "--template",
        type=str,
        help=(
            "Path to section definitions template (CSV or XLSX). "
            "If not specified, uses default template from config.SECTION_TEMPLATE_PATH"
        ),
    )
    parser.add_argument(
        "--use-latest",
        action="store_true",
        help="Use latest available quarter if newer than specified",
    )
    parser.add_argument("--output", type=str, help="Output JSON file path (optional)")

    args = parser.parse_args()

    # Generate execution ID
    execution_id = uuid.uuid4()
    logger.info(f"[START] WM Readthrough ETL v2 (Template-Driven) | execution_id: {execution_id}")

    # Setup context
    ssl_config = setup_ssl()
    auth_config = await setup_authentication(execution_id=str(execution_id), ssl_config=ssl_config)

    context = {
        "execution_id": str(execution_id),
        "ssl_config": ssl_config,
        "auth_config": auth_config,
    }

    # Load section template
    if args.template:
        template_path = args.template
    else:
        template_path = SECTION_TEMPLATE_PATH

    try:
        section_defs = load_section_template(str(template_path))

        # Process all sections
        results = await process_all_sections(
            section_defs=section_defs,
            fiscal_year=args.year,
            quarter=args.quarter,
            context=context,
            use_latest=args.use_latest,
        )

        # Check if any results have content
        has_any_content = any(r.get("has_content") for r in results)

        if not has_any_content:
            logger.warning("[WARNING] No results with content generated")
        else:
            logger.info(
                f"[SUCCESS] Generated {sum(1 for r in results if r.get('has_content'))} "
                f"extractions with content out of {len(results)} total"
            )

        # Determine output path
        if args.output:
            output_path = args.output
        else:
            output_dir = Path(__file__).parent / "output"
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            content_hash = hashlib.md5(
                f"{args.year}_{args.quarter}_{timestamp}".encode()
            ).hexdigest()[:8]
            output_path = (
                output_dir / f"WM_Readthrough_{args.year}_{args.quarter}_{content_hash}.json"
            )

        # Save results
        await save_to_database(
            results=results,
            fiscal_year=args.year,
            quarter=args.quarter,
            execution_id=execution_id,
            output_json_path=str(output_path),
        )

        logger.info("[COMPLETE] WM Readthrough ETL completed successfully")

    except Exception as e:
        logger.error(f"[ERROR] WM Readthrough ETL failed: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    asyncio.run(main())
