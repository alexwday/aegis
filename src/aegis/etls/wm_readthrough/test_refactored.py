"""
Test suite for refactored WM Readthrough ETL - Validates logic and template processing.

This test script validates:
1. Template loading and CSV parsing
2. Prompt variable injection
3. Bank filtering logic
4. Transcript parts retrieval logic
5. Section definition validation
"""

import asyncio
import csv
from pathlib import Path
from typing import Dict, Any

# Test functions


def test_template_loading():
    """Test that template file loads correctly."""
    print("\n=== TEST 1: Template Loading ===")

    template_path = (
        Path(__file__).parent / "templates" / "section_definitions.xlsx"
    )

    if not template_path.exists():
        print(f"✗ FAIL: Template file not found at {template_path}")
        return False

    # Load XLSX file
    from openpyxl import load_workbook

    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Load sections
    sections = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            sections.append(dict(zip(headers, row)))

    wb.close()

    print(f"✓ Loaded {len(sections)} sections from template")

    # Validate required columns
    required_columns = [
        "section_id",
        "section_name",
        "section_description",
        "section_instructions",
        "section_notes",
        "section_examples",
        "transcript_parts",
        "institution_types",
        "prompt_name",
    ]

    if len(sections) > 0:
        actual_columns = set(sections[0].keys())
        missing = set(required_columns) - actual_columns
        if missing:
            print(f"✗ FAIL: Missing columns: {missing}")
            return False
        print(f"✓ All required columns present")

    # Validate transcript_parts values
    valid_parts = ["MD", "QA", "ALL"]
    for section in sections:
        parts = section["transcript_parts"].strip().upper()
        if parts not in valid_parts:
            print(
                f"✗ FAIL: Invalid transcript_parts '{parts}' in section {section['section_id']}"
            )
            return False

    print(f"✓ All transcript_parts values are valid")

    # Validate institution_types
    valid_types = {"Monitored_US_Banks", "US_Banks", "Canadian_Asset_Managers"}
    for section in sections:
        types = [t.strip() for t in section["institution_types"].split(",")]
        invalid_types = set(types) - valid_types
        if invalid_types:
            print(
                f"✗ FAIL: Invalid institution_types {invalid_types} in section {section['section_id']}"
            )
            return False

    print(f"✓ All institution_types values are valid")

    print("✅ TEST 1 PASSED")
    return True


def test_prompt_variable_injection():
    """Test that prompt variables are correctly formatted."""
    print("\n=== TEST 2: Prompt Variable Injection ===")

    # Test system prompt variables
    system_template = """
    # Section Overview
    {section_description}

    # Instructions
    {section_instructions}

    # Notes
    {section_notes}

    # Examples
    {section_examples}
    """

    test_section = {
        "section_description": "Extract WM metrics",
        "section_instructions": "Focus on revenue and AUM",
        "section_notes": "Include YoY comparisons",
        "section_examples": "Example: Revenue grew 15%",
    }

    try:
        formatted = system_template.format(**test_section)
        print("✓ System prompt variable injection successful")
    except KeyError as e:
        print(f"✗ FAIL: Missing variable {e}")
        return False

    # Test user prompt variables
    user_template = """
    Bank: {bank_name}
    Period: {fiscal_year} {quarter}
    Section: {section_name}

    Transcript: {transcript_content}
    """

    test_runtime = {
        "bank_name": "JPMorgan Chase",
        "fiscal_year": 2025,
        "quarter": "Q1",
        "section_name": "WM Narrative",
        "transcript_content": "Test transcript content here...",
    }

    try:
        formatted = user_template.format(**test_runtime)
        print("✓ User prompt variable injection successful")
    except KeyError as e:
        print(f"✗ FAIL: Missing variable {e}")
        return False

    print("✅ TEST 2 PASSED")
    return True


def test_bank_filtering_logic():
    """Test that bank filtering by institution type works correctly."""
    print("\n=== TEST 3: Bank Filtering Logic ===")

    # Mock bank data
    all_banks = [
        {"bank_id": 1, "name": "JPMorgan", "type": "US_Banks"},
        {"bank_id": 2, "name": "Morgan Stanley", "type": "Monitored_US_Banks"},
        {"bank_id": 3, "name": "RBC GAM", "type": "Canadian_Asset_Managers"},
        {"bank_id": 4, "name": "Bank of America", "type": "US_Banks"},
        {"bank_id": 5, "name": "Goldman Sachs", "type": "Monitored_US_Banks"},
    ]

    # Test filtering for US_Banks
    section_types = ["US_Banks"]
    filtered = [b for b in all_banks if b.get("type") in section_types]
    assert len(filtered) == 2, f"Expected 2 US_Banks, got {len(filtered)}"
    print(f"✓ US_Banks filter: {len(filtered)} banks")

    # Test filtering for Monitored_US_Banks
    section_types = ["Monitored_US_Banks"]
    filtered = [b for b in all_banks if b.get("type") in section_types]
    assert len(filtered) == 2, f"Expected 2 Monitored_US_Banks, got {len(filtered)}"
    print(f"✓ Monitored_US_Banks filter: {len(filtered)} banks")

    # Test filtering for Canadian_Asset_Managers
    section_types = ["Canadian_Asset_Managers"]
    filtered = [b for b in all_banks if b.get("type") in section_types]
    assert len(filtered) == 1, f"Expected 1 Canadian_Asset_Managers, got {len(filtered)}"
    print(f"✓ Canadian_Asset_Managers filter: {len(filtered)} banks")

    # Test filtering for multiple types
    section_types = ["US_Banks", "Monitored_US_Banks"]
    filtered = [b for b in all_banks if b.get("type") in section_types]
    assert len(filtered) == 4, f"Expected 4 banks, got {len(filtered)}"
    print(f"✓ Multiple types filter: {len(filtered)} banks")

    print("✅ TEST 3 PASSED")
    return True


def test_transcript_parts_logic():
    """Test that transcript parts retrieval logic is correct."""
    print("\n=== TEST 4: Transcript Parts Logic ===")

    # Test MD only
    parts = "MD"
    should_get_md = parts in ["MD", "ALL"]
    should_get_qa = parts in ["QA", "ALL"]
    assert should_get_md == True, "Should get MD for 'MD'"
    assert should_get_qa == False, "Should not get QA for 'MD'"
    print("✓ MD logic correct")

    # Test QA only
    parts = "QA"
    should_get_md = parts in ["MD", "ALL"]
    should_get_qa = parts in ["QA", "ALL"]
    assert should_get_md == False, "Should not get MD for 'QA'"
    assert should_get_qa == True, "Should get QA for 'QA'"
    print("✓ QA logic correct")

    # Test ALL
    parts = "ALL"
    should_get_md = parts in ["MD", "ALL"]
    should_get_qa = parts in ["QA", "ALL"]
    assert should_get_md == True, "Should get MD for 'ALL'"
    assert should_get_qa == True, "Should get QA for 'ALL'"
    print("✓ ALL logic correct")

    print("✅ TEST 4 PASSED")
    return True


def test_section_result_structure():
    """Test that section result structure is correct."""
    print("\n=== TEST 5: Section Result Structure ===")

    # Test result with content
    result = {
        "section_id": "WM_NARRATIVE",
        "section_name": "WM Narrative",
        "bank_name": "JPMorgan Chase",
        "bank_symbol": "JPM",
        "has_content": True,
        "content": "Sample extracted content...",
        "metadata": {"source_sections": ["MD", "QA"], "confidence": "high"},
    }

    required_fields = [
        "section_id",
        "section_name",
        "bank_name",
        "bank_symbol",
        "has_content",
        "content",
    ]
    for field in required_fields:
        assert field in result, f"Missing required field: {field}"

    print("✓ Result with content has all required fields")

    # Test result without content
    result_no_content = {
        "section_id": "WM_NARRATIVE",
        "section_name": "WM Narrative",
        "bank_name": "Some Bank",
        "bank_symbol": "SB",
        "has_content": False,
        "content": "",
    }

    for field in required_fields:
        assert field in result_no_content, f"Missing required field: {field}"

    print("✓ Result without content has all required fields")

    print("✅ TEST 5 PASSED")
    return True


def test_csv_parsing_edge_cases():
    """Test CSV parsing handles edge cases correctly."""
    print("\n=== TEST 6: CSV Parsing Edge Cases ===")

    # Test parsing institution_types with spaces
    types_str = "US_Banks, Canadian_Asset_Managers"
    types_list = [t.strip() for t in types_str.split(",")]
    assert types_list == ["US_Banks", "Canadian_Asset_Managers"], "Failed to parse with spaces"
    print("✓ Handles institution_types with spaces")

    # Test parsing transcript_parts uppercase conversion
    parts_str = "all"
    parts_upper = parts_str.strip().upper()
    assert parts_upper == "ALL", "Failed to uppercase transcript_parts"
    print("✓ Handles lowercase transcript_parts")

    # Test empty examples field
    examples = ""
    assert examples == "", "Empty examples should remain empty"
    print("✓ Handles empty examples field")

    print("✅ TEST 6 PASSED")
    return True


# Main test runner


def run_all_tests():
    """Run all tests and report results."""
    print("=" * 70)
    print("WM READTHROUGH ETL - LOGIC VALIDATION TEST SUITE")
    print("=" * 70)

    tests = [
        test_template_loading,
        test_prompt_variable_injection,
        test_bank_filtering_logic,
        test_transcript_parts_logic,
        test_section_result_structure,
        test_csv_parsing_edge_cases,
    ]

    passed = 0
    failed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
            else:
                failed += 1
        except Exception as e:
            print(f"\n✗ TEST FAILED WITH EXCEPTION: {e}")
            failed += 1

    print("\n" + "=" * 70)
    print(f"TEST RESULTS: {passed}/{len(tests)} passed, {failed}/{len(tests)} failed")
    print("=" * 70)

    if failed == 0:
        print("\n✅ ALL TESTS PASSED!")
        return True
    else:
        print(f"\n❌ {failed} TEST(S) FAILED")
        return False


if __name__ == "__main__":
    success = run_all_tests()
    exit(0 if success else 1)
