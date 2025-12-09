"""
Script to generate Try a Prompt example questions Excel file for Aegis UI.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_excel():
    """Create the Try a Prompt Excel file with all questions."""
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============ SHEET 1: Landing Page Questions ============
    ws1 = wb.active
    ws1.title = "Landing Page"

    landing_data = [
        ["Database", "Question"],
        ["Supplementary", "What is the FTE count for all Canadian banks in Q3 2025?"],
        ["Supplementary", "Compare the efficiency ratio and ROE for TD and RBC in the latest quarter."],
        ["RTS", "What is RBC's total Level 3 securities balance for Q2 2025?"],
        ["RTS", "What is CIBC's diluted EPS, both reported and adjusted, for Q3 2025?"],
        ["Pillar3", "What is the CET1 ratio for RBC and National Bank in Q2 2025?"],
        ["Pillar3", "What is the leverage ratio for RBC and National Bank in Q2 2025?"],
        ["Transcripts", "What were the key themes from RBC's Q3 2025 earnings call?"],
        ["Transcripts", "Summarize the management discussion from TD's Q2 2025 earnings call."],
    ]

    for row_idx, row_data in enumerate(landing_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 80

    # ============ SHEET 2: Popup Questions ============
    ws2 = wb.create_sheet("Popup Questions")

    popup_data = [
        ["Database", "#", "Question"],
        ["Supplementary", 1, "What is the net income and ROE for US banks in the Capital Markets segment for the latest quarter?"],
        ["Supplementary", 2, "What are the total investment fees and FICC revenue for all US banks in Q3 2025?"],
        ["Supplementary", 3, "What is the ACL for all Canadian banks in Q2 2025?"],
        ["Supplementary", 4, "What is the AUA and net income for RBC's Wealth Management segment in Q3 2025?"],
        ["Supplementary", 5, "What is the AUA and AUM for Scotiabank's Wealth Management Canada segment in YTD 2025?"],
        ["RTS", 1, "What were TD's key corporate events in Q1 2025?"],
        ["RTS", 2, "How many common shares outstanding do the Canadian banks have as of Q2 2025?"],
        ["RTS", 3, "What are TD's assets, deposits, return on RWA, CET1 ratio, and LCR ratio for Q4 2020?"],
        ["RTS", 4, "Summarize BMO's Management Discussion and Analysis for Q2 2025 in five key points."],
        ["RTS", 5, "What does CIBC disclose about its strategy, risks, and outlook in Q3 2024?"],
        ["Pillar3", 1, "What is the CET1 ratio for all Canadian banks across Q1, Q2, and Q3 2025?"],
        ["Pillar3", 2, "What are the Tier 1 ratio and leverage ratio for all Canadian banks in Q2 2025?"],
        ["Pillar3", 3, "What are TD's risk-weighted assets and capital adequacy ratio from their latest Pillar 3 disclosure?"],
        ["Pillar3", 4, "Compare the liquidity coverage ratio and leverage ratio across the Big Six banks for 2025."],
        ["Pillar3", 5, "When was RBC's Pillar 3 disclosure last updated, and what were the significant changes?"],
        ["Transcripts", 1, "What outlook and guidance did RBC management provide in Q3 2025?"],
        ["Transcripts", 2, "What were the main topics analysts asked about in TD's Q2 2025 earnings call?"],
        ["Transcripts", 3, "What strategic initiatives did BMO management highlight in Q3 2025?"],
        ["Transcripts", 4, "Summarize the key points from Scotiabank's Q2 2025 management discussion."],
        ["Transcripts", 5, "Compare the key themes from RBC and TD's Q3 2025 earnings calls."],
    ]

    for row_idx, row_data in enumerate(popup_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 5
    ws2.column_dimensions['C'].width = 90

    # ============ SHEET 3: Dropdown - What is ============
    ws3 = wb.create_sheet("Dropdown - What is")

    what_is_data = [
        ["#", "Question", "Database"],
        [1, "What is RBC's total Level 3 securities balance for Q2 2025?", "RTS"],
        [2, "What is the CET1 ratio for RBC and National Bank in Q2 2025?", "Pillar3"],
        [3, "What is the AUA and net income for RBC's Wealth Management segment in Q3 2025?", "Supplementary"],
        [4, "What is CIBC's diluted EPS, both reported and adjusted, for Q3 2025?", "RTS"],
        [5, "What is the key guidance RBC management provided in Q3 2025?", "Transcripts"],
    ]

    for row_idx, row_data in enumerate(what_is_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 80
    ws3.column_dimensions['C'].width = 15

    # ============ SHEET 4: Dropdown - Compare ============
    ws4 = wb.create_sheet("Dropdown - Compare")

    compare_data = [
        ["#", "Question", "Database"],
        [1, "Compare the efficiency ratio and ROE for TD and RBC in the latest quarter.", "Supplementary"],
        [2, "Compare the management outlook from RBC and TD in Q3 2025.", "Transcripts"],
        [3, "Compare the liquidity coverage ratio and leverage ratio across the Big Six banks for 2025.", "Pillar3"],
        [4, "Compare investment fees and FICC revenue for RBC and BMO in Q3 2025.", "Supplementary"],
        [5, "Compare the key risk factors disclosed by TD and Scotiabank in Q2 2025.", "RTS"],
    ]

    for row_idx, row_data in enumerate(compare_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 80
    ws4.column_dimensions['C'].width = 15

    # ============ SHEET 5: Dropdown - How did ============
    ws5 = wb.create_sheet("Dropdown - How did")

    how_did_data = [
        ["#", "Question", "Database"],
        [1, "How did TD management describe performance and outlook in Q2 2025?", "Transcripts"],
        [2, "How did RBC management respond to analyst concerns in Q3 2025?", "Transcripts"],
        [3, "How did BMO management describe their strategic priorities in Q3 2025?", "Transcripts"],
        [4, "How did TD's risk-weighted assets change from Q1 to Q2 2025?", "Pillar3"],
        [5, "How did CIBC describe their risk management strategy in Q2 2025?", "RTS"],
    ]

    for row_idx, row_data in enumerate(how_did_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws5.column_dimensions['A'].width = 5
    ws5.column_dimensions['B'].width = 80
    ws5.column_dimensions['C'].width = 15

    # ============ SHEET 6: Dropdown - Summarize ============
    ws6 = wb.create_sheet("Dropdown - Summarize")

    summarize_data = [
        ["#", "Question", "Database"],
        [1, "Summarize BMO's Management Discussion and Analysis for Q2 2025 in five key points.", "RTS"],
        [2, "Summarize the key themes from Scotiabank's Q3 2025 earnings call.", "Transcripts"],
        [3, "Summarize TD's capital adequacy metrics from their latest Pillar 3 disclosure.", "Pillar3"],
        [4, "Summarize RBC's strategy, risks, and outlook from Q3 2025 filings.", "RTS"],
        [5, "Summarize the analyst Q&A session from RBC's Q3 2025 earnings call.", "Transcripts"],
    ]

    for row_idx, row_data in enumerate(summarize_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws6.column_dimensions['A'].width = 5
    ws6.column_dimensions['B'].width = 80
    ws6.column_dimensions['C'].width = 15

    # Save workbook
    output_path = "/Users/alexwday/Projects/aegis/aegis_try_a_prompt_questions.xlsx"
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    create_excel()
